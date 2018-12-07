'''
Created on Nov 24, 2018
@author: micha and trinh
'''
#Importing packages used
import openpyxl
import numpy as np
import xlsxwriter
from macpath import realpath


#Importing data from "Line_Data.xlsx" excel document
workbook = openpyxl.load_workbook('454 Data Project.xlsx') 

MVAbase = 100

#Making each sheet of data (Line & Load data) more accessable
lineData = workbook['LineData'] 
busData = workbook['BusData'] 

#Initializing arrays to store Pload,Qload,V,Bus Type
#Where we assume bus 1 is the "Slack Bus" 
Pload = [] 
Qload = [] 
Pgen = [] 
V = [] 
busType = [] 

#Tracking largest active and reactive power mismatch
largestPmis = np.zeros((1,2))
largestQmis = np.zeros((1,2))

#Putting BusData into arrays
for row_index in range(1, (busData.max_row)): 
    Pload.append(busData.cell(row=row_index + 1, column=2).value/MVAbase)
    Qload.append(busData.cell(row=row_index + 1, column=3).value/MVAbase)
    Pgen.append(busData.cell(row=row_index + 1, column=4).value/MVAbase)
    V.append(busData.cell(row=row_index + 1, column=5).value)
    busType.append(busData.cell(row=row_index + 1, column=6).value)
    
# print('Pload:',Pload)
# print('Qload:',Qload) 
# print('Pgen:', Pgen)
# print('V:', V)
# print('busType:', busType)

# print("~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~")

#Initializing Y-Admittitance matrix where the size is based on the 
#number of buses in the grid
numBuses = 12
yBus = np.zeros(shape=(numBuses, numBuses), dtype=complex)

for row_index in range (0, (lineData.max_row - 1)): 
    
    sendIndex = lineData.cell(row=row_index + 2, column=1).value
    receiveIndex = lineData.cell(row_index + 2, column=2).value
    
    rTotal = lineData.cell(row_index + 2, column=3).value
    xTotal = lineData.cell(row_index + 2, column=4).value
    bTotal = lineData.cell(row_index + 2, column=5).value
    seriesImpedance = complex(rTotal, xTotal)
    
    #Populating off-diagonal terms
    yBus[sendIndex - 1, receiveIndex - 1] = -1*(1 / seriesImpedance)
    yBus[receiveIndex - 1, sendIndex - 1] = -1*(1 / seriesImpedance)

    #Populating diagondal terms
    yBus[sendIndex - 1, sendIndex - 1] += (1 / seriesImpedance) + (1j * bTotal / 2)
    yBus[receiveIndex - 1, receiveIndex - 1] += (1 / seriesImpedance) + (1j * bTotal / 2)

#print(yBus)
#a = np.asarray(yBus)
#np.savetxt("foo.csv", a, delimiter=",")

# print("~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~")

#Convergence requirements
Pconv = 0.1 / MVAbase
Qconv = 0.1 / MVAbase

#Number of PQ and PV buses
numPV = 0
numPQ = 0
#Injections have all known P injections first followed by known Q injections

#Count how many PQ and PV buses in system
for idx in range(0,len(busType)):
    if busType[idx] == "PV":
        numPV += 1

    elif busType[idx] == "PQ":
        numPQ += 1


#fill unknown V's with flat start guesses
for idx in range(0, numBuses):
    if busType[idx] =="PQ":
        V[idx] = 1

#intialize a theta array with knowns and flat start guesses
theta = np.zeros(numBuses)

#Function to identify indexes of PQ buses
def pqIdx():
    pqIdx = []
    for idx in range(0, numBuses):
        if busType[idx] =="PQ":
            pqIdx.append(idx+1)
    return pqIdx

pqIdx = pqIdx()


#Function for P computed
def Pcomp(j):
    Pcomp = 0
    for k in range(0, numBuses):
        Pcomp += V[j] * V[k] * (np.real(yBus[j][k])*np.cos(theta[j]-theta[k]) + np.imag(yBus[j][k])*np.sin(theta[j]-theta[k]))
    return Pcomp

#Function for Q computed
def Qcomp(i):
    Qcomp = 0
    for k in range(0, numBuses):
        Qcomp += V[i] * V[k] * (np.real(yBus[i][k])*np.sin(theta[i]-theta[k]) - np.imag(yBus[i][k])*np.cos(theta[i]-theta[k]))
    return Qcomp

#Function for P mismatch
def deltaP():
    deltaP = np.zeros(numBuses-1)
    for j in range(0, numBuses - 1):
        deltaP[j] = Pcomp(j+1) - (Pgen[j+1]-Pload[j+1])
    return deltaP

#Function for Q mismatch
def deltaQ():
    deltaQ = np.zeros(numPQ)
    for j in range(0, numPQ):
        deltaQ[j] = (Qcomp(pqIdx[j]-1) - (-1*Qload[pqIdx[j]-1]))
    return deltaQ


#Define Jacobian submatrices
def J11():
    J11 = np.zeros(shape = (numBuses-1, numBuses-1))

    for j in range(0,numBuses-1):
        for k in range(0,numBuses-1):
            if (k) == (j):
                J11[j][k] = -1*Qcomp(j+1) -1*V[j+1]**2*np.imag(yBus[j+1][j+1])
            else:
                J11[j][k] = V[j+1] * V[k+1] *(np.real(yBus[j+1][k+1])*np.sin(theta[j+1]-theta[k+1]) - np.imag(yBus[j+1][k+1])*np.cos(theta[j+1]-theta[k+1]))
    return J11


def J12():
    J12 = np.zeros(shape = (numBuses-1, numPQ))

    for j in range(0,numBuses-1):
        for k in range (0,numPQ):
            kprime = pqIdx[k]
            if (kprime-1) == (j+1):
                J12[j][k] = Pcomp(j+1)/V[j+1] + V[j+1]*np.real(yBus[j+1][j+1])
            else:
                J12[j][k] = V[j+1] *(np.real(yBus[j+1][kprime-1])*np.cos(theta[j+1]-theta[kprime-1]) + np.imag(yBus[j+1][kprime-1])*np.sin(theta[j+1]-theta[kprime-1]))
    return J12

def J21():
    J21 = np.zeros(shape=(numPQ, numBuses-1))
    
    for j in range(0,numPQ):
        kprime = pqIdx[j] 
        for k in range(0,numBuses-1): 
            if kprime-1 == k+1: 
                J21[j][k] = Pcomp(kprime-1) - (np.real(yBus[kprime-1][kprime-1])*V[kprime-1]**2)
            else: 
                J21[j][k] = ((-1) * V[kprime-1] * V[k+1]) * ((np.real(yBus[kprime-1][k+1]) * np.cos(theta[kprime-1] - theta[k+1])) + 
                                                          (np.imag(yBus[kprime-1][k+1]) * np.sin(theta[kprime-1] - theta[k+1])))

    return J21

def J22():
    J22 = np.zeros(shape=(numPQ, numPQ))
    
    for j in range(0,numPQ): 
        kprime1 = pqIdx[j]
        for k in range(0,numPQ): 
            kprime2 = pqIdx[k]
            if kprime1-1 == kprime2-1: 
                J22[j][k] = (Qcomp(kprime1-1)/V[kprime1-1])-(np.imag(yBus[kprime1-1][kprime1-1])*V[kprime1-1])
            else:
                J22[j][k] = V[kprime1-1]*((np.real(yBus[kprime1-1][kprime2-1])*np.sin(theta[kprime1-1]-theta[kprime2-1])) -
                                       (np.imag(yBus[kprime1-1][kprime2-1])*np.cos(theta[kprime1-1]-theta[kprime2-1]))) 
    return J22

def J():
    J1 = np.vstack((J11(),J21()))
    J2 = np.vstack((J12(),J22()))
    J = np.hstack((J1,J2))

    return J

def calcCorrection(jacobian, mismatch):
    invJacobian = np.linalg.inv(jacobian)
    correction = np.matmul(-1*invJacobian, mismatch)
    return correction

def update(jacobian, mismatch, numBuses):
    for idx in range(0,(numBuses-1)):
        theta[idx+1] += calcCorrection(jacobian, mismatch)[idx]
    
    for idx in range(0,(numPQ)):
        kprime = pqIdx[idx]
        V[kprime-1] += calcCorrection(jacobian, mismatch)[numBuses-1+ idx][0]
    return;

def checkConverge(dP, dQ):
    for p in range(0, len(dP)): 
        if dP[p] > Pconv: 
            #print(p)
            return False;
    for q in range(0, len(dQ)): 
        if dQ[q] > Qconv: 
            return False;
    return True;

def findLargeMis(dP, dQ):
    global largestPmis
    global largestQmis
    
    if abs(np.min(dP)) > abs(largestPmis[0][0]) or np.max(dP) > abs(largestPmis[0][0]): 
        if abs(np.min(dP)) > np.max(dP): 
            largestPmis[0][0] = np.min(dP)
            largestPmis[0][1] = np.argmin(dP)+1
        else:
            largestPmis[0][0] = np.max(dP)
            largestPmis[0][1] = np.argmax(dP)+1
            
    if abs(np.min(dQ)) > abs(largestQmis[0][0]) or np.max(dQ) > abs(largestQmis[0][0]): 
        if abs(np.min(dQ)) > np.max(dQ): 
            largestQmis[0][0] = np.min(dQ)
            largestQmis[0][1] = pqIdx[np.argmin(dQ)]
        else:
            largestQmis[0][0] = np.max(dQ)
            largestQmis[0][1] = pqIdx[np.argmax(dQ)]
    return;
def pFlow(send,receive):
    pFlow = (V[send]*V[receive])*(-1*yBus[send][receive])*np.sin(theta[send]-theta[receive])
    return abs(pFlow)

def qFlow(send,receive):
    qFlow = (V[send]*V[receive])*(-1*yBus[send][receive])*np.cos(theta[send]-theta[receive])-((V[receive])**2*(-1*yBus[send][receive]))
    
    return abs(qFlow)

result_Pgen = []
result_PgenInd = []
result_Qgen = []
result_QgenInd = []

def PandGresults(numBuses):
    for idx in range(0,numBuses):
        if busType[idx] == "S" or busType[idx] == "PV":
            result_Pgen.append(str((Pcomp(idx)+Pload[idx])*MVAbase))
            result_PgenInd.append(str(idx+1))
            result_Qgen.append(((Qcomp(idx)+Qload[idx])*MVAbase))
            result_QgenInd.append(str(idx+1))
    return;

sendInd = [] 
receiveInd = [] 
realP = [] 
reacP = [] 
limit = []
def RealandReacP():
    for idx in range(0, (lineData.max_row - 1)):
        send = lineData.cell(idx + 2, column=1).value
        receive = lineData.cell(idx + 2, column=2).value
        fMax = lineData.cell(idx + 2, column=6).value
        
        sendInd.append(send)
        receiveInd.append(receive)
        realP.append(pFlow(send-1,receive-1)*MVAbase)
        reacP.append(qFlow(send-1,receive-1)*MVAbase)
    
        if (pFlow(send-1,receive-1)*MVAbase > fMax) or (qFlow(send-1,receive-1)*MVAbase > fMax):
            limit.append('Limit Exceeded')
        else:
            limit.append('Within MVA limit') 
    return;

def WriteExcel1(col, param):
    row = 0
    for ind in (param): 
        worksheet1.write(row+1, col, ind)
        row += 1    
    return;

def WriteExcel2(col, param):
    row = 0
    for ind in (param): 
        worksheet2.write(row+1, col, ind)
        row += 1    
    return;

def WriteBusData(numBuses, theta, V): 
    worksheet1.write(0,0,'Bus #')
    worksheet1.write(0,1,'Voltage (p.u.)')
    worksheet1.write(0,2,'Angle (degrees)')
    worksheet1.write(0,3,'Largest P Mismatch')
    worksheet1.write(0,4,'Largest P Mismatch Bus')
    worksheet1.write(0,5,'Largest Q Mismatch')
    worksheet1.write(0,6,'Largest Q Mismatch Bus')
    
    row = 0
    for ind in range(0, numBuses):
        worksheet1.write(row+1,0,ind+1)
        row += 1
        
    row = 0  
    for angle in (theta): 
        worksheet1.write(row+1, 2, angle * 57.2958)
        row += 1
        
    WriteExcel1(1, V)
    
    worksheet1.write(1,3,largestPmis[0][0])
    worksheet1.write(1,4,largestPmis[0][1])
    worksheet1.write(1,5,largestQmis[0][0])
    worksheet1.write(1,6,largestQmis[0][1])    
    
    return;

def WriteLineData():
    worksheet2.write(0,0,'Bus #')
    worksheet2.write(0,1,'P_gen (MW)')
    worksheet2.write(0,3,'Send Bus')
    worksheet2.write(0,4,'Receiving Bus')
    worksheet2.write(0,5,'Real Power (MW)')
    worksheet2.write(0,6,'Reactive Power (MVAr)')
    worksheet2.write(0,7,'Limit Warning MVAr')
    
    PandGresults(numBuses)
    RealandReacP()
    
    WriteExcel2(0, result_PgenInd)
    WriteExcel2(1, result_Pgen)
    WriteExcel2(3, sendInd)
    WriteExcel2(4, receiveInd)
    WriteExcel2(5, realP)
    WriteExcel2(6, reacP)
    WriteExcel2(7, limit)
    return;

if __name__ == '__main__':
    for i in range(0,25):
        print("Number of Iterations:", i)
        P = deltaP()
        Q = deltaQ()
        
        dP = P.reshape(len(P),1)
        dQ = Q.reshape(len(Q),1)
        findLargeMis(P, Q)
        
        mismatch = np.vstack((dP,dQ))
        
        if checkConverge(deltaP(),deltaQ()) == True: 
            print(mismatch)
            print("theta")
            print(theta)
            print("V")
            print(V)
            print("Largest P Mismatch:", largestPmis)
            print("Largest Q Mismatch:", largestQmis)
            
            break
        
        update(J(), mismatch, numBuses)
        ###################################################################
        
    workbook = xlsxwriter.Workbook('Results.xlsx')
    worksheet1 = workbook.add_worksheet('busData')
    worksheet2 = workbook.add_worksheet('lineData')
    
    WriteBusData(numBuses, theta, V)
    WriteLineData()
    
    workbook.close()
    pass

