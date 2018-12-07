'''
Created on Nov 24, 2018
@author: micha and trinh
'''
#Importing packages used
import openpyxl
import numpy as np
import math
import cmath

#Parameters to specify
MVAbase = 100
numBuses = 5
Pconv = 0.1
Qconv = 0.1

#Specifies excel document we are working with: "Line_Data.xlsx" 
workbook = openpyxl.load_workbook('454 Data.xlsx') 

#Making the sheets inside excel document (Line & Load data) more accessable
lineData = workbook['LineData'] 
busData = workbook['BusData'] 

#Initializing arrays to store Pload, Qload, Pgen, V, Bus Type
#Where we assume bus 1 is the "Slack Bus" 
Pload = [] 
Qload = [] 
Pgen = [] 
V = [] 
busType = [] 

#Intialize a theta array with knowns and flat start guesses
theta = np.zeros(numBuses)

#Injections have all known P injections first followed by known Q injections
inj = []

#Function uses "busData" which specifies the excel sheet
#used to take and assign values into the Pload, Qload, Pgen, V, and 
#busType arrays
def importBusData():
    #Putting BusData into arrays where each column corresponds to the
    #arrays defined in the beginning
    
    for row_index in range(1, (busData.max_row)): 
        Pload.append(busData.cell(row=row_index + 1, column=2).value/MVAbase)
        Qload.append(busData.cell(row=row_index + 1, column=3).value/MVAbase)
        Pgen.append(busData.cell(row=row_index + 1, column=4).value/MVAbase)
        V.append(busData.cell(row=row_index + 1, column=5).value)
        busType.append(busData.cell(row=row_index + 1, column=6).value)
    return; 

importBusData() 

#Counter for PQ and PV buses
numPV = 0
numPQ = 0

#Count how many PQ and PV buses in system and appends delta P in the
#beginning of inj (injection) array and delta Q in the end of the array
def countBuses():

    for idx in range(0,len(busType)):
        
        global numPV
        global numPQ
        
        #Appends P to beginning of array
        if busType[idx] == "PV":
            numPV += 1
            inj.append(Pgen[idx] - Pload[idx])
            
        elif busType[idx] == "PQ":
            numPQ += 1
            inj.append(Pgen[idx] - Pload[idx])
    
    #Appends Q to the end of the array
    for idx in range(0,len(busType)):
        if busType[idx] == "PQ":
            inj.append(-1*Qload[idx])
    return;

countBuses()

#For explicit Q equation
V_pqIndex = []

#Fill unknown V's with flat start guesses
for idx in range(0, numBuses):
    if busType[idx] =="PQ":
        V_pqIndex.append(idx + 1)
        V[idx] = 1
"""
print(V)
print(V_pqIndex)
print(V[11])
print(len(V_pqIndex)) 
print(V_pqIndex[0])
print(V_pqIndex[6])
"""
#Initializing Y-Admittitance matrix where the size is based on the 
#number of buses in the grid, which is specified in the beginning
#of the program
yBus = np.zeros(shape=(numBuses, numBuses), dtype=complex)

#Function uses "lineData" which specifies the excel sheet 
#used to take and assign values into the Y-Bus and converts any 
# Rtotal and Xtotal values into G's and B's
def importLineData():
    
    for row_index in range (0, (lineData.max_row - 1)): 
        
        sendIndex = lineData.cell(row=row_index + 2, column=1).value
        receiveIndex = lineData.cell(row_index + 2, column=2).value
        
        rTotal = lineData.cell(row_index + 2, column=3).value
        xTotal = lineData.cell(row_index + 2, column=4).value
        bTotal = lineData.cell(row_index + 2, column=5).value
        seriesImpedance = complex(rTotal, xTotal)
        
        #Populating off-diagonal terms (Rtotal and Xtotal values)
        yBus[sendIndex - 1, receiveIndex - 1] = -1*(1 / seriesImpedance)
        yBus[receiveIndex - 1, sendIndex - 1] = -1*(1 / seriesImpedance)
    
        #Populating diagonal terms (sum of Rtotal, Xtotal, and Btotal)
        yBus[sendIndex - 1, sendIndex - 1] += (1 / seriesImpedance) + (1j * bTotal / 2)
        yBus[receiveIndex - 1, receiveIndex - 1] += (1 / seriesImpedance) + (1j * bTotal / 2)
    return;

importLineData()

#print("~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~")

#Mismatch equations
deltaP = np.zeros(shape = numBuses - 1)
deltaQ = np.zeros(shape = numPQ)
Pcomp = np.zeros(shape = numBuses - 1)
Qcomp = np.zeros(shape = numBuses - 1)
#Qcomp = np.zeros(shape = numPQ)
"""
print(deltaP)
print(Pload)
print(theta)
print(V)
print(yBus)
print("~~~~~~~~~~~~~~~~~~~~~~~~")
"""
def Pmismatch(numBuses):
    for k in range(0, numBuses-1):
        for i in range(0, numBuses):
            #print("i:", i)
            #print("k", k)
            Pcomp[k] += (V[k + 1]*V[i])*((np.real(yBus[k + 1][i])*np.cos(theta[k + 1]-theta[i])) + 
                                          (np.imag(yBus[k + 1][i])*np.sin(theta[k + 1]-theta[i])))
        
        deltaP[k] = Pcomp[k] - (Pgen[k+1] - Pload[k+1])
    return;
Pmismatch(numBuses) 
"""
def Qcomp(numBuses):
    for k in range (0, numBuses-1): 
        for i in range(0, numBuses):
            Qcomp[k] += V[k + 1]*V[i]*(np.real(yBus[k + 1][i])*np.sin(theta[k + 1]-theta[i]) - np.imag(yBus[k + 1][i])*np.cos(theta[k + 1]-theta[i]))
                                                      
    return;
"""
def Qmismatch(numPQ, numBuses):
    for k in range (0, numPQ): 
        for i in range(0, numBuses):
            index = V_pqIndex[k]-1 
            Qcomp[k] += V[index]*V[i] * (np.real(yBus[index][i])*np.sin(theta[index]-theta[i]) - 
                                          np.imag(yBus[index][i])*np.cos(theta[index]-theta[i]))

        deltaQ[k] = -Qcomp[k] - Qload[index] 
    return;

Qmismatch(numPQ, numBuses)
#print(deltaQ)


#Consider creating a theta matrix, in which all values are initially 0
#Consider modifying the V matrix, in which all zero values are 1
#Define Jacobian submatrices
def J11(numBuses):
    J11 = np.zeros(shape = (numBuses-1, numBuses-1))

    for j in range(0,numBuses-1):
        for k in range (0,numBuses-1):
            if j+1 == k+1: 
                J11[j][k] = ((-1)*Qcomp[j])-(V[j+1]**2 * np.real(yBus[j+1][j+1]))
            else:
                J11[j][k] = V[j+1] * V[k+1] *(np.real(yBus[j+1][k+1])*np.sin(theta[j+1]-theta[k+1]) - 
                                              np.imag(yBus[j+1][k+1])*np.cos(theta[j+1]-theta[k+1]))
    return J11


#print(J11(numBuses))
#print("~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~")

def J12(numBuses, numPQ):
    
    J12 = np.zeros(shape = (numBuses-1, numPQ))

    for j in range(0,numBuses-1):
        for k in range (0,numPQ):
            
            index = V_pqIndex[k]-1
            
            if j+1 == k+1: 
                J12[j][k] = (Pcomp[j]/V[j+1]) + (np.real(yBus[j+1][j+1])*V[j+1])
            else: 
                J12[j][k] = V[j+1] *(np.real(yBus[j+1][index])*np.cos(theta[j+1]-theta[index]) + 
                                     np.imag(yBus[j+1][index])*np.sin(theta[j+1]-theta[index]))
    return J12;

#print(J12(numBuses, numPQ))
#print("~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~")

def J21(numBuses, numPQ):
    J21 = np.zeros(shape=(numPQ, numBuses-1))
    
    for j in range(0,numPQ): 
        for k in range(0,numBuses-1): 
            index = V_pqIndex[j]-1
            if j+1 == k+1: 
                J21[j][k] = Pcomp[j] - (np.real(yBus[j+1][j+1])*V[j+1]**2)
            else: 
                J21[j][k] = ((-1) * V[index] * V[k+1]) * ((np.real(yBus[index][k+1]) * np.cos(theta[index] - theta[k+1])) + 
                                                          (np.imag(yBus[index][k+1]) * np.sin(theta[index] - theta[k+1])))

    return J21;

#print(J21(numBuses, numPQ))
#print("~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~")

def J22():
    J22 = np.zeros(shape=(numPQ, numPQ))
    
    for j in range(0,numPQ): 
        for k in range(0,numPQ): 
            index = V_pqIndex[j]-1
            
            if j+1 == k+1: 
                J22[j][j] = (Qcomp[j]/V[j+1])-(np.real(yBus[j+1][j+1])*V[j+1])
            else:
                J22[j][k] = V[index]*((np.real(yBus[index][k+1])*np.sin(theta[index]-theta[k+1])) -
                                       (np.imag(yBus[index][k+1])*np.cos(theta[index]-theta[k+1]))) 
    
    return J22

def checkConverge(deltaP, deltaQ):
    convergence = 1 
    for p in range(0, len(deltaP)): 
        if deltaP[p] > Pconv: 
            return False;
    if convergence == 1: 
        for q in range(0, len(deltaQ)): 
            if deltaQ[q] > Qconv: 
                return False;
        if convergence == 1: 
            return True; 
    return False;

print(J22())
print("~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~")

print(Qcomp)
def update(Vdiff, Thetadiff):
    global V
    global theta 
    return;

jacobian = np.full((18,18), 1)
jacobian = np.linalg.inv(-jacobian)
print(jacobian)  

# a = J11(numBuses, numPQ)
# np.savetxt("foo.csv", a, delimiter=",")

# a = np.asarray(yBus)
# np.savetxt("foo.csv", a, delimiter=",")
#print(deltaP)
#print(deltaQ)
#print(V)
"""
print(Pload)
print(Qload)
print(Pgen)
print(V)
print(busType)
print(numPV)
print(numPQ)
"""