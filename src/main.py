'''
Created on Nov 24, 2018

@author: micha and trinh
'''
#Importing packages used
import openpyxl
import numpy as np
import math
import cmath

#Importing data from "Line_Data.xlsx" excel document
workbook = openpyxl.load_workbook('454 Data.xlsx') 
print(workbook)


#Making each sheet of data (Line & Load data) more accessable
lineData = workbook['LineData'] 
busData = workbook['BusData'] 

#Initializing arrays to store P,Q,V,Bus Type
#Where we assume bus 1 is the "Slack Bus" 
P = [] 
Q = [] 
Pgen = [] 
V = [] 
busType = [] 

#Putting BusData into arrays where there is nested 
#forloop to go row by row, and column by column 
#when counter = 1 put into P array 
#             = 2          Q array 
#             = 3          Pgen array 
#             = 4          V array 
#             = 5          busType array 
MVAbase = 100 
for row_index in range(1, (busData.max_row)): 
    P.append(busData.cell(row=row_index + 1, column=2).value / MVAbase)
    Q.append(busData.cell(row=row_index + 1, column=3).value / MVAbase)
    Pgen.append(busData.cell(row=row_index + 1, column=4).value / MVAbase)
    V.append(busData.cell(row=row_index + 1, column=5).value)
    busType.append(busData.cell(row=row_index + 1, column=6).value)
print('P:',P)
print('Q:',Q) 
print('Pgen:', Pgen)
print('V:', V)
print('busType:', busType)

print("~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~")
#Initializing Y-Admittitance matrix where the size is based on the 
#number of buses in the grid
numberOfBuses = 12
yAdmittance = np.zeros(shape=(numberOfBuses, numberOfBuses), dtype=complex)
print(yAdmittance)
print("~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~")

for row_index in range (0, (lineData.max_row - 1)): 
    
    sendIndex = lineData.cell(row=row_index + 2, column=1).value
    receiveIndex = lineData.cell(row_index + 2, column=2).value
    
    rTotal = lineData.cell(row_index + 2, column=3).value
    xTotal = lineData.cell(row_index + 2, column=4).value
    bTotal = lineData.cell(row_index + 2, column=5).value
    seriesImpedance = complex(rTotal, xTotal)
    
    yAdmittance[sendIndex - 1, receiveIndex - 1] = -1*(1 / seriesImpedance)
    yAdmittance[receiveIndex - 1, sendIndex - 1] = -1*(1 / seriesImpedance)
    
    yAdmittance[sendIndex - 1, sendIndex - 1] += (1 / seriesImpedance) + (1j * bTotal / 2)
    yAdmittance[receiveIndex - 1, receiveIndex - 1] += (1 / seriesImpedance) + (1j * bTotal / 2)

print(yAdmittance)
a = np.asarray(yAdmittance)
np.savetxt("foo.csv", a, delimiter=",")

print("~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~")
#Convergence requirements
Pconv = 0.1
Qconv = 0.1

for idx in range(0,len(busType)):
    if busType[idx] == "S":
        print("S")
    elif busType[idx] == "PV":
        print("PV")
    elif busType[idx] == "PQ":
        P_inj = Pgen[idx] - P[idx]
        print(P_inj)
