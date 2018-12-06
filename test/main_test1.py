'''
Created on Nov 24, 2018

@author: micha and trinh
'''
#Importing packages used
import openpyxl
import numpy as np
import math
import cmath

#Importing data from "Line_Data.xlsx" excel document
workbook = openpyxl.load_workbook('454 Data.xlsx') 

MVAbase = 100

#Making each sheet of data (Line & Load data) more accessable
lineData = workbook['LineData'] 
busData = workbook['BusData'] 

#Initializing arrays to store Pload,Qload,V,Bus Type
#Where we assume bus 1 is the "Slack Bus" 
Pload = [] 
Qload = [] 
Pgen = [] 
V = [] 
busType = [] 

#Putting BusData into arrays
for row_index in range(1, (busData.max_row)): 
    Pload.append(busData.cell(row=row_index + 1, column=2).value/MVAbase)
    Qload.append(busData.cell(row=row_index + 1, column=3).value/MVAbase)
    Pgen.append(busData.cell(row=row_index + 1, column=4).value/MVAbase)
    V.append(busData.cell(row=row_index + 1, column=5).value)
    busType.append(busData.cell(row=row_index + 1, column=6).value)
    
# print('Pload:',Pload)
# print('Qload:',Qload) 
# print('Pgen:', Pgen)
# print('V:', V)
# print('busType:', busType)

# print("~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~")

#Initializing Y-Admittitance matrix where the size is based on the 
#number of buses in the grid
numBuses = 5
yBus = np.zeros(shape=(numBuses, numBuses), dtype=complex)

for row_index in range (0, (lineData.max_row - 1)): 
    
    sendIndex = lineData.cell(row=row_index + 2, column=1).value
    receiveIndex = lineData.cell(row_index + 2, column=2).value
    
    rTotal = lineData.cell(row_index + 2, column=3).value
    xTotal = lineData.cell(row_index + 2, column=4).value
    bTotal = lineData.cell(row_index + 2, column=5).value
    seriesImpedance = complex(rTotal, xTotal)
    
    #Populating off-diagonal terms
    yBus[sendIndex - 1, receiveIndex - 1] = -1*(1 / seriesImpedance)
    yBus[receiveIndex - 1, sendIndex - 1] = -1*(1 / seriesImpedance)

    #Populating diagondal terms
    yBus[sendIndex - 1, sendIndex - 1] += (1 / seriesImpedance) + (1j * bTotal / 2)
    yBus[receiveIndex - 1, receiveIndex - 1] += (1 / seriesImpedance) + (1j * bTotal / 2)

print(yBus)
a = np.asarray(yBus)
np.savetxt("foo.csv", a, delimiter=",")

# print("~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~")

#Convergence requirements
Pconv = 0.1
Qconv = 0.1

#Number of PQ and PV buses
numPV = 0
numPQ = 0
#Injections have all known P injections first followed by known Q injections

#Count how many PQ and PV buses in system
for idx in range(0,len(busType)):
    if busType[idx] == "PV":
        numPV += 1

    elif busType[idx] == "PQ":
        numPQ += 1


#fill unknown V's with flat start guesses
for idx in range(0, numBuses):
    if busType[idx] =="PQ":
        V[idx] = 1

#intialize a theta array with knowns and flat start guesses
theta = np.zeros(numBuses)


#Function to identify indexes of PQ buses
def pqIdx(numPQ):
    pqIdx = []
    for idx in range(0, numBuses):
        if busType[idx] =="PQ":
            pqIdx.append(idx+1)
    return pqIdx

pqIdx = pqIdx(numPQ)

#Function for P computed
def Pcomp(numBuses,j):
    Pcomp = np.zeros(numBuses-1)
    for k in range(0, numBuses):

        Pcomp[j] += V[j+1] * V[k] * (np.real(yBus[j+1][k])*np.cos(theta[j+1]-theta[k]) + np.imag(yBus[j+1][k])*np.sin(theta[j+1]-theta[k]))

    return Pcomp
#Function for Q computed
def Qcomp(numPQ,j,i):
    Qcomp = np.zeros(numPQ)

    for k in range(0, numBuses):

        Qcomp[j] += V[i-1] * V[k] * (np.real(yBus[i-1][k])*np.sin(theta[i-1]-theta[k]) - np.imag(yBus[i-1][k])*np.cos(theta[i-1]-theta[k]))
    return Qcomp

#Function for P mismatch
def deltaP(numBuses):
    deltaP = np.zeros(numBuses-1)

    for j in range(0, numBuses - 1):
        deltaP[j] = (Pgen[j+1]-Pload[j+1]) - Pcomp(numBuses,j)[j]

    return deltaP
#Function for Q mismatch
def deltaQ(numPQ):
    deltaQ = np.zeros(numPQ)
    for j in range(0, numPQ):

        deltaQ[j] = (-1*Qload[pqIdx[j]-1]) - Qcomp(numPQ,j,pqIdx[j])[j]
    return deltaQ


#Define Jacobian submatrices
def J11(numBuses):
    J11 = np.zeros(shape = (numBuses-1, numBuses-1))

    for j in range(0,numBuses-1):
        for k in range(0,numBuses-1):
            
            if (k+1) == (j+1):
                for l in range(0,numBuses):
                    J11[j][k] = -1*Qcomp(numBuses,j,l)[j] -1*V[j+1]**2*np.imag(yBus[j+1][j+1])
            else:
                J11[j][k] = V[j+1] * V[k+1] *(np.real(yBus[j+1][k+1])*np.sin(theta[j+1]-theta[k+1]) - np.imag(yBus[j+1][k+1])*np.cos(theta[j+1]-theta[k+1]))

    return J11

def J12(numBuses, numPQ):
    J12 = np.zeros(shape = (numBuses-1, numPQ))

    for j in range(0,numBuses-1):
        for k in range (0,numPQ):
            kprime = pqIdx[k]
            if (kprime+1) == (j+1):

                    J12[j][k] = (Pcomp(numBuses,j)[j]/V[kprime-1]) + np.real(yBus[j+1][kprime-1])*V[kprime-1]
            else:
                J12[j][k] = V[j+1] *(np.real(yBus[j+1][pqIdx[k]-1])*np.cos(theta[j+1]-theta[pqIdx[k]-1]) + np.imag(yBus[j+1][pqIdx[k]-1])*np.sin(theta[j+1]-theta[pqIdx[k]-1]))

    return J12

def J21(numBuses, numPQ):
    J21 = np.zeros(shape = (numPQ, numBuses - 1))
    pqIdx = [2,3,4]



    for j in range(0,numPQ):
        for k in range (0,numBuses - 1):

            J21[j][k] = V[pqIdx[j]]*V[k] *(np.real(yBus[j+1][k+1])*np.cos(theta[j+1]-theta[k+1]) + np.imag(yBus[j+1][k+1])*np.sin(theta[j+1]-theta[k+1]))

    return J21

#print(J11(5))

a = np.asarray(J12(numBuses,numPQ))
np.savetxt("foo.csv", a, delimiter=",")

def J(J11,J12,J21,J22):
    J1 = np.vstack((J11,J21))
    J2 = np.vstack((J12,J22))
    J = np.hstack((J1,J2))

    return J

#Calculate mismatch = [deltaP deltaQ]
#If mismatch is within limits, done abs(mismatch) < 0.1
#Build Jacobian
#corrections = -J[x0]^-1*Mismatch[x0]
#update: [delta[x1], v[x1]] = [delta[x0] v[x0]] + corrections

#repeat above steps until mismatches converge